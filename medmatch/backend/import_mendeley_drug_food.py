"""Import the CC BY 4.0 Mendeley Drug-Food evidence workbook."""
from __future__ import annotations

import sqlite3
from pathlib import Path

from openpyxl import load_workbook

from .db import DB_PATH

DATA_PATH = Path(__file__).parent / "data" / "mendeley_drug_food.xlsx"
SOURCE = "Mendeley Drug-Food Interactions (CC BY 4.0)"
SCHEMA = """
CREATE TABLE IF NOT EXISTS mendeley_drug_food_2021 (
    label INTEGER NOT NULL,
    food_constituent TEXT NOT NULL,
    food_smiles TEXT,
    drug_constituent TEXT NOT NULL,
    drug_smiles TEXT,
    interaction TEXT NOT NULL,
    source TEXT NOT NULL,
    PRIMARY KEY (food_constituent, drug_constituent, interaction)
);
"""


def import_rows(conn: sqlite3.Connection) -> int:
    conn.executescript(SCHEMA)
    conn.execute("DELETE FROM mendeley_drug_food_2021")
    wb = load_workbook(DATA_PATH, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        label, food, food_smiles, drug, drug_smiles, interaction = row[:6]
        if food and drug and interaction:
            rows.append((int(label or 0), str(food).strip(), food_smiles,
                         str(drug).strip(), drug_smiles, str(interaction).strip(), SOURCE))
    conn.executemany(
        "INSERT OR IGNORE INTO mendeley_drug_food_2021"
        " (label, food_constituent, food_smiles, drug_constituent, drug_smiles, interaction, source)"
        " VALUES (?,?,?,?,?,?,?)", rows)
    conn.commit()
    wb.close()
    return conn.execute("SELECT COUNT(*) FROM mendeley_drug_food_2021").fetchone()[0]


if __name__ == "__main__":
    conn = sqlite3.connect(DB_PATH)
    try:
        print(f"imported {import_rows(conn)} Mendeley Drug-Food rows")
    finally:
        conn.close()
